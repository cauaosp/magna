import openpyxl
import json
import schedule
from datetime import datetime, timedelta
from scapy.all import IP, ICMP, sr1, conf, get_if_addr
import requests
from types import SimpleNamespace
import json

def load_json_as_object():
    with open("auth/auth.json", "r") as file:
        return json.load(file, object_hook=lambda d: SimpleNamespace(**d))

auth = load_json_as_object()

def read_sheet(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    excel_list = []

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        linha_dict = {}
        for i, valor in enumerate(row):
            if valor is not None:
                linha_dict[header[i]] = valor
        if linha_dict.get("IP") is not None:
            excel_list.append(linha_dict)
    return excel_list

def load_email_config():
    with open('auth/auth.json', 'r') as f:
        data = json.load(f)
    
    email_config = data["email_smtp"]
    return email_config

def html_content(broke_printers):
    table = """
    <html>
        <head>
            <meta http-equiv='Content-Type' content='text/html; charset=UTF-8'>
        </head>
        <body>
            <p><strong>Problemas com as seguintes impressoras:</strong></p>
            <table style="border-collapse: collapse; table-layout: fixed; width: 100%;">
                <tr>
                    <th style="border: 1px solid #ddd; padding: 8px; background-color: #f2f2f2; text-align: center; width: 18rem">Nome da Impressora</th>
                    <th style="border: 1px solid #ddd; padding: 8px; background-color: #f2f2f2; text-align: center; width: 6rem">IP</th>
                    <th style="border: 1px solid #ddd; padding: 8px; background-color: #f2f2f2; text-align: center; width: 6rem">Horário</th>
                    <th style="border: 1px solid #ddd; padding: 8px; background-color: #f2f2f2; text-align: center;">Problema</th>
                </tr>
    """

    for printer in broke_printers:
        table += f"""
        <tr style="background-color: #ffffff">
            <td style="border: 1px solid #ddd; padding: 8px; text-align: left;">{printer.get('Nome da Impressora')}</td>
            <td style="border: 1px solid #ddd; padding: 8px; text-align: left;">{printer.get('IP')}</td>
            <td style="border: 1px solid #ddd; padding: 8px; text-align: left;">{printer.get('Horario')}</td>
            <td style="border: 1px solid #ddd; padding: 8px; text-align: left; white-space: pre-wrap; font-family: Arial, sans-serif;"><pre>{printer.get('Problema')}</pre></td>
        </tr>
        """

    table += """
            </table>
        </body>
    </html>
    """

    return table

def handle_ping(printer, time, broke_printers):
    sent = 0
    received = 0
    lost = 0

    for _ in range(2):
        packet = IP(dst=printer.get("IP")) / ICMP() / b"Ping personalizado!"
        response = sr1(packet, timeout=2, verbose=0)
        sent += 1
        if response:
            received += 1
        else:
            lost += 1

    if lost == 0:
        if printer in broke_printers:
            broke_printers.remove(printer)
    else:
        if printer not in broke_printers:
            printer["Horario"] = time
            printer["Problema"] = f"Estatísticas do Ping de {get_if_addr(conf.iface.name)}  para {printer.get("IP")}:\nPacotes: Enviados = {sent}, Recebidos = {received}, Perdidos = {lost}\n{(lost / sent) * 100}% de perda"
            broke_printers.append(printer)

def schedule_ping_for_printers(excel_list, broke_printers):
    hourly_ticket_times = []
    current_time = datetime.strptime("06:00", "%H:%M")
    end_time = datetime.strptime("23:59", "%H:%M")
    
    while current_time <= end_time:
        hourly_ticket_times.append(current_time.time())
        current_time += timedelta(hours=1)

    for printer in excel_list:
        start_time = printer["Horario Inicial"]
        end_time = printer["Horario Final"]

        if isinstance(start_time, datetime):
            start_dt = start_time
        else:
            start_dt = datetime.combine(datetime.today(), start_time)

        if isinstance(end_time, datetime):
            end_dt = end_time
        else:
            end_dt = datetime.combine(datetime.today(), end_time)

        current_time = start_dt

        while current_time <= end_dt:
            time_str = current_time.strftime("%H:%M")
            schedule.every().day.at(time_str).do(handle_ping, printer=printer, broke_printers=broke_printers, time=time_str)
            print(f"Agendado para {printer['Nome da Impressora']} às {time_str}")
            current_time += timedelta(hours=1)

    for time in hourly_ticket_times:
        time_str = time.strftime("%H:%M")
        schedule.every().day.at(time_str).do(lambda printers=broke_printers: handle_ticket_creation(printers))
        print(f"Agendamento para criar ticket às {time_str}")

def get_session_token():
    session = requests.Session()
    session.auth = (auth.glpi_api.user.username, auth.glpi_api.user.password)
    response = session.get(auth.glpi_api.requests.newSession, headers={"Content-Type": "application/json", "App-Token": auth.glpi_api.appToken, "User-Token": auth.glpi_api.userToken})
    return response.json()["session_token"]

def end_session_token(token):
    session = requests.Session()
    session.auth = (auth.glpi_api.user.username, auth.glpi_api.user.password)
    response = session.get(auth.glpi_api.requests.endSession, headers={"Content-Type": "application/json", "App-Token": auth.glpi_api.appToken, "User-Token": auth.glpi_api.userToken, "Session-Token": token})
    print(response.json())
    if(response.json() == True):
        print("Token de sessão encerrado!")
    else:
        print("Token inválido para encerrar!")

def create_new_ticket(body):
    session = requests.Session()
    session.auth = (auth.glpi_api.user.username, auth.glpi_api.user.password)
    token = get_session_token()
    respoonse = session.post(auth.glpi_api.requests.newTicket, headers={"Content-Type": "application/json", "App-Token": auth.glpi_api.appToken, "Session-Token": token}, json={
        "input": {
            "name": "PingBot - Impressoras quebradas",
            "content": html_content(body),
            "status": 2,
            "itilcategories_id": 128,
            "locations_id": 393,
            "entities_id": 1
        }
    })
    print(respoonse.status_code)
    print("Ticket criado!")
    end_session_token(token)

def is_ticket_created():
    session = requests.session()
    session.auth = (auth.glpi_api.user.username, auth.glpi_api.user.password)
    token = get_session_token()
    respoonse = session.get(auth.glpi_api.requests.search + "?criteria[0][field]=7&criteria[0][searchtype]=equals&criteria[0][value]=128&criteria[1][link]=AND&criteria[1][field]=1&criteria[1][searchtype]=contains&criteria[1][value]=PingBot", headers={"Content-Type": "application/json", "App-Token": auth.glpi_api.appToken, "Session-Token": token})
    data = respoonse.json().get("data")
    end_session_token(token)

    if(data):
        print("JÁ TEM CHAMADO, ADICIONA TAREFA NOVA!")
        return True
    else:
        print("NÂO HÁ CHAMADO, CRIE UM NOVO!")
        return False


def handle_ticket_creation(broke_printers):
    if(len(broke_printers) == 0):
        print("Não há impressoras quebradas!")
    else:
        print(f"Total de impressoras quebradas: {len(broke_printers)}")
        if(is_ticket_created()):
            create_task(broke_printers)
        else:
            create_new_ticket(broke_printers)

def create_task(broke_printers):
    print("\nProblemas com estas impressoras\n")
    for printer in broke_printers:
        nome = printer['Nome da Impressora']
        ip = printer['IP']
        horario = printer['Horario']
        problema = printer['Problema']
        print(f"{nome}\t{ip}\t{horario}")
        print(problema + "\n")