import openpyxl 

wb = openpyxl.load_workbook('arquivo.xlsx')

if __name__ == "__main__":
    ws = wb.worksheets[0]
    ws.title = "Arquivo Testado"
    print(ws.title)
    print(f"Dados da planilha {wb.sheetnames}\n")

    my_list = list()

    ws['A2'] = "Primeiro dia"
    ws.cell(row=1, column=4, value="RESULTADO")

    for value in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=4, values_only=True):
        my_list.append(value)
        
    for ele1,ele2,ele3, ele4 in my_list:
        print(f"{str(ele1 or '-'):<15}{str(ele2 or '-'):<15}{str(ele3 or '-'):<15}{str(ele4 or '-'):<15}")
    print(f"total de linhas: {ws.max_row}")


    # print("\nPlanilhas adicionada:")
    # wb.create_sheet('impressoras')
    # print(wb.sheetnames)
    
    # print("\nPlanilha excluída, total:")
    # wb.remove(wb['planilha 3'])
    # print(wb.sheetnames)
 
    print(my_list)

    wb.save('arquivo.xlsx')